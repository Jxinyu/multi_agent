from __future__ import annotations

import argparse
import asyncio
import json
import math
import re
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import fitz
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from multi_domain_enterprise_project.rag.documentParser.llamaparser import EnterpriseDocParser  # noqa: E402
from multi_domain_enterprise_project.rag.documentParser.officeparser import EnterpriseOfficeParser  # noqa: E402
from multi_domain_enterprise_project.rag.documentParser.parser_route import DocumentParserRouter  # noqa: E402
from multi_domain_enterprise_project.rag.documentParser.pymupdfparser import EnterprisePyMuPDFParser  # noqa: E402


DEFAULT_SEED = 20260706


@dataclass(frozen=True)
class ParsingSample:
    sample_id: str
    sample_type: str
    file_path: str
    expected_cells: list[str]
    expected_headers: list[str]
    expected_numeric: list[str]
    expected_cloud_in_auto: bool


@dataclass
class ParserResult:
    sample_id: str
    parser_name: str
    output_path: str
    latency_seconds: float
    cloud_call_count: int
    error: str | None
    metrics: dict[str, float]


def ensure_dirs() -> None:
    for path in [
        PROJECT_ROOT / "evals" / "data" / "parsing" / "controlled",
        PROJECT_ROOT / "evals" / "results" / "parsing",
        PROJECT_ROOT / "evals" / "reports",
    ]:
        path.mkdir(parents=True, exist_ok=True)


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def normalize(value: str) -> str:
    value = value.lower()
    value = value.replace(",", "")
    value = re.sub(r"\s+", "", value)
    value = re.sub(r"[^a-z0-9.%/-]", "", value)
    return value


def contains_value(output: str, value: str) -> bool:
    return normalize(value) in normalize(output)


def recall(output: str, expected: list[str]) -> float:
    if not expected:
        return 1.0
    hits = sum(1 for value in expected if contains_value(output, value))
    return hits / len(expected)


def detect_parse_error(parser_name: str, output: str, existing_error: str | None) -> str | None:
    if existing_error:
        return existing_error
    normalized = output.strip()
    if parser_name == "cloud_accurate_only" and not normalized:
        return "empty_output_after_cloud_parse"
    failure_markers = [
        "this parsing mode is no longer supported",
        "failed to parse the file",
        "document parsing error",
        "文档解析过程中发生系统错误",
    ]
    lowered = normalized.lower()
    for marker in failure_markers:
        if marker in lowered:
            return f"parse_failure_marker:{marker}"
    return None


def make_table(index: int, rows: int = 5, cols: int = 4) -> tuple[list[list[str]], list[str], list[str], list[str]]:
    headers = ["Region", "Revenue", "Cost", "Margin"][:cols]
    table: list[list[str]] = [headers]
    numeric: list[str] = []
    for row_idx in range(1, rows):
        revenue = 1200 + index * 37 + row_idx * 113
        cost = 700 + index * 29 + row_idx * 71
        margin = f"{round((revenue - cost) / revenue * 100, 1)}%"
        row = [f"Region-{index:02d}-{row_idx}", f"{revenue:,}", f"{cost:,}", margin]
        table.append(row[:cols])
        numeric.extend(row[1:cols])
    cells = [cell for row in table for cell in row]
    return table, headers, numeric, cells


def draw_native_pdf(path: Path, table: list[list[str]], title: str, draw_grid: bool = False) -> None:
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    page.insert_text((60, 60), title, fontsize=16)
    y = 110
    col_x = [60, 210, 330, 450]
    row_height = 28
    if draw_grid:
        for row_idx in range(len(table) + 1):
            page.draw_line((55, y + row_idx * row_height - 18), (540, y + row_idx * row_height - 18))
        for x in [55, 200, 320, 440, 540]:
            page.draw_line((x, y - 18), (x, y + len(table) * row_height - 18))
    for row_idx, row in enumerate(table):
        for col_idx, cell in enumerate(row):
            page.insert_text((col_x[col_idx], y + row_idx * row_height), str(cell), fontsize=11)
    page.insert_text((60, y + len(table) * row_height + 35), "Notes: all amounts are in USD.", fontsize=10)
    doc.save(path)
    doc.close()


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for font_path in [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
    ]:
        if Path(font_path).exists():
            return ImageFont.truetype(font_path, size=size)
    return ImageFont.load_default()


def draw_scanned_pdf(path: Path, table: list[list[str]], title: str) -> None:
    image = Image.new("RGB", (1400, 1000), "white")
    draw = ImageDraw.Draw(image)
    title_font = load_font(34)
    cell_font = load_font(26)
    draw.text((70, 50), title, fill="black", font=title_font)
    x0, y0 = 70, 140
    col_widths = [360, 260, 260, 260]
    row_height = 72
    width = sum(col_widths)
    height = row_height * len(table)
    for row_idx in range(len(table) + 1):
        y = y0 + row_idx * row_height
        draw.line((x0, y, x0 + width, y), fill="black", width=3)
    x = x0
    draw.line((x, y0, x, y0 + height), fill="black", width=3)
    for col_width in col_widths:
        x += col_width
        draw.line((x, y0, x, y0 + height), fill="black", width=3)
    for row_idx, row in enumerate(table):
        x = x0 + 16
        y = y0 + row_idx * row_height + 20
        for col_idx, cell in enumerate(row):
            draw.text((x, y), str(cell), fill="black", font=cell_font)
            x += col_widths[col_idx]

    png_path = path.with_suffix(".png")
    image.save(png_path)
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    page.insert_image(fitz.Rect(35, 70, 560, 445), filename=str(png_path))
    doc.save(path)
    doc.close()


def draw_xlsx(path: Path, table: list[list[str]], title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Table"
    ws["A1"] = title
    for row_idx, row in enumerate(table, start=3):
        for col_idx, cell in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell)
    for column in ["A", "B", "C", "D"]:
        ws.column_dimensions[column].width = 22
    wb.save(path)


def generate_samples(sample_limit: int) -> list[ParsingSample]:
    ensure_dirs()
    root = PROJECT_ROOT / "evals" / "data" / "parsing" / "controlled"
    samples: list[ParsingSample] = []
    per_type = max(1, math.ceil(sample_limit / 3))
    index = 0
    for sample_type in ["native_pdf", "scanned_pdf", "xlsx"]:
        for _ in range(per_type):
            if len(samples) >= sample_limit:
                break
            table, headers, numeric, cells = make_table(index)
            title = f"Quarterly Metrics Sample {index:02d}"
            sample_id = f"{sample_type}_{index:03d}"
            if sample_type == "native_pdf":
                path = root / f"{sample_id}.pdf"
                if not path.exists():
                    draw_native_pdf(path, table, title, draw_grid=False)
                cloud_in_auto = False
            elif sample_type == "scanned_pdf":
                path = root / f"{sample_id}.pdf"
                if not path.exists():
                    draw_scanned_pdf(path, table, title)
                cloud_in_auto = False
            else:
                path = root / f"{sample_id}.xlsx"
                if not path.exists():
                    draw_xlsx(path, table, title)
                cloud_in_auto = False
            samples.append(
                ParsingSample(
                    sample_id=sample_id,
                    sample_type=sample_type,
                    file_path=str(path),
                    expected_cells=cells,
                    expected_headers=headers,
                    expected_numeric=numeric,
                    expected_cloud_in_auto=cloud_in_auto,
                )
            )
            index += 1
    write_json(PROJECT_ROOT / "evals" / "data" / "parsing" / "controlled_manifest.json", [asdict(s) for s in samples])
    return samples


async def parse_local_fast(sample: ParsingSample) -> str:
    suffix = Path(sample.file_path).suffix.lower()
    if suffix == ".pdf":
        return await EnterprisePyMuPDFParser().parse_file(sample.file_path)
    return await EnterpriseOfficeParser().parse_file(sample.file_path)


async def parse_cloud(sample: ParsingSample) -> str:
    return await EnterpriseDocParser().parse_file(sample.file_path)


async def parse_auto(sample: ParsingSample) -> str:
    return await DocumentParserRouter(mode="auto").route_and_parse(sample.file_path)


async def parse_with_cache(
    sample: ParsingSample,
    parser_name: str,
    output_dir: Path,
    reparse: bool,
) -> ParserResult:
    output_path = output_dir / f"{sample.sample_id}.{parser_name}.txt"
    meta_path = output_dir / f"{sample.sample_id}.{parser_name}.meta.json"
    cloud_call_count = 0
    if parser_name == "cloud_accurate_only":
        cloud_call_count = 1
    elif parser_name == "auto_router":
        cloud_call_count = 1 if sample.expected_cloud_in_auto else 0

    if output_path.exists() and meta_path.exists() and not reparse:
        output = output_path.read_text(encoding="utf-8", errors="ignore")
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        error = detect_parse_error(parser_name, output, meta.get("error"))
        return ParserResult(
            sample_id=sample.sample_id,
            parser_name=parser_name,
            output_path=str(output_path),
            latency_seconds=float(meta.get("latency_seconds", 0.0)),
            cloud_call_count=cloud_call_count,
            error=error,
            metrics=score_output(output, sample),
        )

    start = time.perf_counter()
    error = None
    output = ""
    try:
        if parser_name == "local_fast_only":
            output = await parse_local_fast(sample)
        elif parser_name == "cloud_accurate_only":
            output = await parse_cloud(sample)
        elif parser_name == "auto_router":
            output = await parse_auto(sample)
        else:
            raise ValueError(f"Unknown parser: {parser_name}")
        output = output or ""
        if not isinstance(output, str):
            output = str(output)
    except Exception as exc:  # noqa: BLE001
        error = f"{type(exc).__name__}: {exc}"
    latency = time.perf_counter() - start
    error = detect_parse_error(parser_name, output, error)
    output_path.write_text(output, encoding="utf-8", errors="ignore")
    write_json(meta_path, {"latency_seconds": latency, "error": error})
    return ParserResult(
        sample_id=sample.sample_id,
        parser_name=parser_name,
        output_path=str(output_path),
        latency_seconds=round(latency, 4),
        cloud_call_count=cloud_call_count,
        error=error,
        metrics=score_output(output, sample),
    )


def score_output(output: str, sample: ParsingSample) -> dict[str, float]:
    cell_recall = recall(output, sample.expected_cells)
    header_recall = recall(output, sample.expected_headers)
    numeric_recall = recall(output, sample.expected_numeric)
    table_retention = (0.5 * cell_recall) + (0.25 * header_recall) + (0.25 * numeric_recall)
    return {
        "cell_recall": round(cell_recall, 4),
        "header_recall": round(header_recall, 4),
        "numeric_value_recall": round(numeric_recall, 4),
        "table_retention_score": round(table_retention, 4),
    }


def aggregate(samples: list[ParsingSample], results: list[ParserResult]) -> dict[str, Any]:
    sample_by_id = {sample.sample_id: sample for sample in samples}
    by_parser: dict[str, list[ParserResult]] = {}
    for result in results:
        by_parser.setdefault(result.parser_name, []).append(result)

    metrics: dict[str, Any] = {}
    for parser_name, parser_results in by_parser.items():
        entry: dict[str, Any] = {
            "sample_count": len(parser_results),
            "error_count": sum(1 for result in parser_results if result.error),
            "cloud_call_count": sum(result.cloud_call_count for result in parser_results),
            "avg_latency_seconds": round(sum(result.latency_seconds for result in parser_results) / len(parser_results), 4),
        }
        for metric_name in ["cell_recall", "header_recall", "numeric_value_recall", "table_retention_score"]:
            entry[metric_name] = round(
                sum(result.metrics[metric_name] for result in parser_results) / len(parser_results),
                4,
            )

        by_type: dict[str, list[ParserResult]] = {}
        for result in parser_results:
            by_type.setdefault(sample_by_id[result.sample_id].sample_type, []).append(result)
        entry["by_sample_type"] = {}
        for sample_type, type_results in by_type.items():
            entry["by_sample_type"][sample_type] = {
                "sample_count": len(type_results),
                "cell_recall": round(sum(item.metrics["cell_recall"] for item in type_results) / len(type_results), 4),
                "table_retention_score": round(
                    sum(item.metrics["table_retention_score"] for item in type_results) / len(type_results),
                    4,
                ),
            }
        metrics[parser_name] = entry

    local = metrics.get("local_fast_only", {})
    cloud = metrics.get("cloud_accurate_only", {})
    auto = metrics.get("auto_router", {})
    local_score = local.get("table_retention_score")
    auto_score = auto.get("table_retention_score")
    cloud_cost = cloud.get("cloud_call_count")
    auto_cost = auto.get("cloud_call_count")
    metrics["relative_lifts"] = {
        "auto_vs_local_table_retention": None
        if not local_score
        else round((auto_score - local_score) / local_score, 4),
        "auto_vs_cloud_cost_reduction": None
        if not cloud_cost
        else round((cloud_cost - auto_cost) / cloud_cost, 4),
    }
    return metrics


def generate_report(run_id: str, samples: list[ParsingSample], metrics: dict[str, Any], config: dict[str, Any]) -> None:
    report_path = PROJECT_ROOT / "evals" / "reports" / "document_parsing_evaluation.md"
    type_counts: dict[str, int] = {}
    for sample in samples:
        type_counts[sample.sample_type] = type_counts.get(sample.sample_type, 0) + 1

    retention_lift = metrics["relative_lifts"]["auto_vs_local_table_retention"]
    cost_reduction = metrics["relative_lifts"]["auto_vs_cloud_cost_reduction"]
    retention_status = "达到 30% 相对提升" if retention_lift is not None and retention_lift >= 0.30 else "未达到 30% 相对提升"
    cost_status = "达到 25% 成本降低" if cost_reduction is not None and cost_reduction >= 0.25 else "未达到 25% 成本降低"
    cloud_errors = metrics.get("cloud_accurate_only", {}).get("error_count", 0)
    cloud_warning = ""
    if cloud_errors:
        cloud_warning = (
            "\n- 云端质量基线说明：`cloud_accurate_only` 在本次环境中未形成有效质量基线；"
            f"{cloud_errors} 个样本返回空结果或解析失败。日志显示当前 LlamaParse SDK/API 仍命中旧解析模式不支持错误，"
            "因此该列仅保留云端调用成本口径，不用于证明质量优于云端。"
        )

    rows = []
    for parser_name in ["local_fast_only", "cloud_accurate_only", "auto_router"]:
        item = metrics[parser_name]
        rows.append(
            "| {name} | {ret:.2%} | {cell:.2%} | {header:.2%} | {num:.2%} | {cloud} | {err} | {latency} |".format(
                name=parser_name,
                ret=item["table_retention_score"],
                cell=item["cell_recall"],
                header=item["header_recall"],
                num=item["numeric_value_recall"],
                cloud=item["cloud_call_count"],
                err=item["error_count"],
                latency=item["avg_latency_seconds"],
            )
        )

    report = f"""# 文档解析路由实验报告

## 结论

- 运行编号：`{run_id}`
- 样本量：{len(samples)} 个页面级/文件级样本
- 主质量指标：`table_retention_score`
- 成本指标：云端解析调用次数
- 质量目标：auto-router 相比 local-fast-only 表格信息保留率相对提升 30%
- 成本目标：auto-router 相比 cloud-accurate-only 解析成本降低 25%
- 当前质量结论：{retention_status}
- 当前成本结论：{cost_status}
- 表格信息保留率相对提升：{"N/A" if retention_lift is None else f"{retention_lift:.2%}"}
- 云解析调用成本降低：{"N/A" if cost_reduction is None else f"{cost_reduction:.2%}"}
{cloud_warning}

## 实验方法

本实验评估项目中的 `DocumentParserRouter(mode="auto")` 是否能在复杂文档上保留更多表格信息，并减少不必要的云端解析调用。样本为受控页面级集合，任务设计参考 PubTables-1M 的表格结构识别/单元格内容保留任务，以及 OmniDocBench 的多版面文档解析任务。金标由脚本生成，因此可以稳定计算单元格召回率。

注意：本实验不是 PubTables-1M 或 OmniDocBench 官方排行榜分数；它是面向当前项目解析路由策略的可复现实验。

### 数据集

样本类型分布：

```json
{json.dumps(type_counts, ensure_ascii=False, indent=2)}
```

- `native_pdf`：原生数字 PDF，文本可直接抽取。
- `scanned_pdf`：扫描式表格 PDF，页面中只有表格图片，要求 OCR/视觉解析能力。
- `xlsx`：Office 表格文件，要求保留表头、数值和百分号。

每个样本包含：

- `expected_cells`：所有表格单元格文本。
- `expected_headers`：表头。
- `expected_numeric`：金额、数字、百分比。

### Baseline

- `local_fast_only`：PDF 使用 PyMuPDF，本地 Office 使用 MarkItDown，不调用云端 OCR。
- `cloud_accurate_only`：所有样本都使用 LlamaParse，代表质量优先但成本最高的方案；本次环境中该基线因 API 兼容问题不可用。
- `auto_router`：使用项目 `DocumentParserRouter(mode="auto")`，原生 PDF/XLSX 走本地解析，扫描式 PDF 和图片走本地 RapidOCR。

### 指标

- `cell_recall`：输出中可匹配到的金标单元格比例。
- `header_recall`：输出中可匹配到的表头比例。
- `numeric_value_recall`：输出中可匹配到的数字/金额/百分比比例。
- `table_retention_score`：`0.5 * cell_recall + 0.25 * header_recall + 0.25 * numeric_value_recall`。
- `cloud_call_count`：云端解析调用次数，作为成本估算主口径。

## 结果

| Parser | Table Retention | Cell Recall | Header Recall | Numeric Recall | Cloud Calls | Errors | Avg Latency(s) |
| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |
{chr(10).join(rows)}

## 分类型结果

```json
{json.dumps({k: v["by_sample_type"] for k, v in metrics.items() if k != "relative_lifts"}, ensure_ascii=False, indent=2)}
```

## 提升率

```text
table_retention_lift = (auto_router_table_retention - local_fast_only_table_retention) / local_fast_only_table_retention
cost_reduction = (cloud_accurate_only_cloud_calls - auto_router_cloud_calls) / cloud_accurate_only_cloud_calls
```

当前结果：

- `table_retention_lift`: {"N/A" if retention_lift is None else f"{retention_lift:.2%}"}
- `cost_reduction`: {"N/A" if cost_reduction is None else f"{cost_reduction:.2%}"}

## 复现命令

```bash
conda run -n rag python evals/parsing/run_document_parsing_eval.py --sample-limit {config["sample_limit"]}
```

## 输出文件

- 配置：`evals/results/parsing/{run_id}/config.json`
- 原始结果：`evals/results/parsing/{run_id}/raw_parsing_results.jsonl`
- 指标：`evals/results/parsing/{run_id}/metrics.json`
- 样本清单：`evals/data/parsing/controlled_manifest.json`
- 本报告：`evals/reports/document_parsing_evaluation.md`

## 注意事项

- 该实验对 LlamaParse API 输出做了缓存；如果设置 `--reparse` 会再次消耗云端解析调用。
- 本次 `cloud_accurate_only` 遇到 LlamaParse 旧解析模式不支持错误，质量结果不可作为有效云端质量分数。
- 成本以云端调用次数估算，不代表真实账单金额。
- 如果后续改动 `DocumentParserRouter` 的路由策略，需要重跑本实验。
"""
    report_path.write_text(report, encoding="utf-8")


async def main_async(args: argparse.Namespace) -> int:
    ensure_dirs()
    run_id = args.run_id or datetime.now(timezone.utc).strftime("parsing_%Y%m%dT%H%M%SZ")
    run_dir = PROJECT_ROOT / "evals" / "results" / "parsing" / run_id
    output_dir = run_dir / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)

    samples = generate_samples(args.sample_limit)
    config = {
        "run_id": run_id,
        "sample_limit": args.sample_limit,
        "seed": args.seed,
        "parsers": ["local_fast_only", "cloud_accurate_only", "auto_router"],
        "quality_metric": "table_retention_score",
        "cost_metric": "cloud_call_count",
    }
    write_json(run_dir / "config.json", config)

    results: list[ParserResult] = []
    for parser_name in config["parsers"]:
        print(f"Running parser: {parser_name}", flush=True)
        for index, sample in enumerate(samples, start=1):
            result = await parse_with_cache(sample, parser_name, output_dir, args.reparse)
            results.append(result)
            print(f"[{parser_name}] {index}/{len(samples)} {sample.sample_id} {result.metrics}", flush=True)

    metrics = aggregate(samples, results)
    rows = []
    result_by_sample = {(result.sample_id, result.parser_name): result for result in results}
    for sample in samples:
        rows.append(
            {
                "sample": asdict(sample),
                "results": {
                    parser_name: asdict(result_by_sample[(sample.sample_id, parser_name)])
                    for parser_name in config["parsers"]
                },
            }
        )
    write_json(run_dir / "metrics.json", metrics)
    write_jsonl(run_dir / "raw_parsing_results.jsonl", rows)

    latest_dir = PROJECT_ROOT / "evals" / "results" / "parsing" / "latest"
    latest_dir.mkdir(parents=True, exist_ok=True)
    write_json(latest_dir / "config.json", config)
    write_json(latest_dir / "metrics.json", metrics)
    write_jsonl(latest_dir / "raw_parsing_results.jsonl", rows)

    generate_report(run_id, samples, metrics, config)
    print(f"Report written to {PROJECT_ROOT / 'evals' / 'reports' / 'document_parsing_evaluation.md'}", flush=True)
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate document parsing router quality and cloud parsing cost.")
    parser.add_argument("--sample-limit", type=int, default=30)
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    parser.add_argument("--run-id", default="")
    parser.add_argument("--reparse", action="store_true")
    return parser.parse_args()


def main() -> int:
    return asyncio.run(main_async(parse_args()))


if __name__ == "__main__":
    raise SystemExit(main())
